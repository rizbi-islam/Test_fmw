"""
Creates a BLANK Excel test suite template — no demo data.
Use this as the starting point for your real project tests.

Usage:
    python assets/templates/blank_template.py

Output:
    assets/templates/blank_test_suite.xlsx
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_blank_template(output_path: str = "assets/templates/blank_test_suite.xlsx"):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = openpyxl.Workbook()

    hdr_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    tip_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    tip_font = Font(color="1565C0", italic=True, size=9)
    thin     = Side(style="thin", color="DDDDDD")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    center   = Alignment(horizontal="center", vertical="center")
    left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── TestSuite Sheet ───────────────────────────────────────
    ws = wb.active
    ws.title = "TestSuite"

    headers = [
        "TestCaseID", "TestName", "Enabled", "StepNo", "Keyword",
        "Locator", "Strategy", "Value", "Timeout", "DataSource", "Description"
    ]

    # Header row
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = center; c.border = border
    ws.row_dimensions[1].height = 22

    # Tip rows (row 2 — instructions, greyed out)
    tips = [
        "e.g. TC001", "e.g. Login_Test", "YES/NO", "1,2,3...",
        "NAVIGATE / TYPE / CLICK etc.", "#id or [name='x'] or //xpath",
        "css/xpath/id/name", "URL or text or {variable}", "10",
        "excel/sqlite/json", "Optional step description"
    ]
    for ci, tip in enumerate(tips, 1):
        c = ws.cell(row=2, column=ci, value=tip)
        c.fill = tip_fill; c.font = tip_font
        c.alignment = left; c.border = border
    ws.row_dimensions[2].height = 18

    # 20 empty data rows
    for ri in range(3, 23):
        for ci in range(1, 12):
            c = ws.cell(row=ri, column=ci, value="")
            c.border = border
            c.fill   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") \
                       if ri % 2 == 0 else \
                       PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
        ws.row_dimensions[ri].height = 18

    col_widths = [12, 24, 8, 7, 18, 30, 10, 36, 8, 14, 30]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    ws.freeze_panes = "A3"

    # ── TestData Sheet ────────────────────────────────────────
    ws2 = wb.create_sheet("TestData")
    data_headers = [
        "RowID", "username", "password", "email",
        "phone", "otp", "expected_text", "base_url"
    ]
    for ci, h in enumerate(data_headers, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = center; c.border = border
    ws2.row_dimensions[1].height = 22

    # Tip row
    data_tips = [
        "row1/row2...", "testuser", "Pass@123",
        "user@example.com", "01700000000", "auto/123456",
        "Welcome", "https://your-app.com"
    ]
    for ci, tip in enumerate(data_tips, 1):
        c = ws2.cell(row=2, column=ci, value=tip)
        c.fill = tip_fill; c.font = tip_font
        c.alignment = left; c.border = border

    for ri in range(3, 13):
        for ci in range(1, 9):
            c = ws2.cell(row=ri, column=ci, value="")
            c.border = border
        ws2.row_dimensions[ri].height = 18

    for ci, w in enumerate([10, 16, 14, 26, 14, 10, 20, 30], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    ws2.freeze_panes = "A3"

    # ── Keywords Reference Sheet ──────────────────────────────
    ws3 = wb.create_sheet("Keywords Reference")
    kw_headers = ["Keyword", "Required Params", "Optional Params", "Example Value", "Description"]
    for ci, h in enumerate(kw_headers, 1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = center; c.border = border

    keywords = [
        ("NAVIGATE",       "value",          "—",                "https://app.com",     "Navigate to URL"),
        ("TYPE",           "locator, value", "strategy, timeout","#email — {email}",    "Type text into field"),
        ("CLICK",          "locator",        "strategy, timeout","#submit-btn",         "Click element"),
        ("CLEAR",          "locator",        "strategy",         "#input-field",        "Clear input field"),
        ("SELECT",         "locator, value", "strategy, select_by","#dropdown — Option1","Select dropdown option"),
        ("WAIT_FOR",       "locator",        "strategy, timeout","#loading-spinner",    "Wait until element visible"),
        ("ASSERT_TEXT",    "locator, value", "strategy",         "h1 — Welcome",        "Assert element text"),
        ("ASSERT_VISIBLE", "locator",        "strategy, timeout","#success-msg",        "Assert element exists"),
        ("ASSERT_URL",     "value",          "—",                "dashboard",           "Assert URL contains text"),
        ("VERIFY_URL",     "value",          "timeout",          "/home",               "Wait + verify URL"),
        ("SCREENSHOT",     "—",             "value",            "step1.png",           "Take screenshot"),
        ("SCROLL_TO",      "—",             "locator or value", "bottom",              "Scroll to element/direction"),
        ("HOVER",          "locator",        "strategy",         "#menu-item",          "Hover mouse over element"),
        ("GET_TEXT",       "locator",        "strategy, value",  "#greeting — msg_var", "Get text into variable"),
        ("SLEEP",          "value",          "—",                "2",                   "Pause N seconds"),
        ("CLOSE_BROWSER",  "—",             "—",                "—",                   "Close browser"),
        ("SWITCH_FRAME",   "locator",        "strategy",         "#iframe-id",          "Switch into iframe"),
        ("SWITCH_DEFAULT", "—",             "—",                "—",                   "Switch back to main page"),
        ("HANDLE_OTP",     "locator",        "value, timeout",   "#otp — mock",         "Get OTP and type it"),
        ("HANDLE_CAPTCHA", "—",             "value",            "bypass",              "Solve/bypass CAPTCHA"),
    ]
    for ri, row in enumerate(keywords, 2):
        fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") \
               if ri % 2 == 0 else \
               PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
        for ci, val in enumerate(row, 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.fill = fill; c.border = border
            c.alignment = left
        ws3.row_dimensions[ri].height = 18

    for ci, w in enumerate([18, 22, 22, 30, 40], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws3.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\n  Blank template created: {output_path}")
    print(f"  Sheets: TestSuite (blank), TestData (blank), Keywords Reference")
    print(f"\n  Instructions:")
    print(f"    1. Open in Excel")
    print(f"    2. Fill TestData sheet with your test data")
    print(f"    3. Fill TestSuite sheet with your test steps")
    print(f"    4. Run: python run.py --list")
    print(f"    5. Run: python run.py --headless\n")


if __name__ == "__main__":
    create_blank_template()
