"""
Creates the Excel test suite template with sample test cases.
Run once to generate the working template file.

Usage:
    python assets/templates/create_template.py
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_template(output_path: str = "assets/templates/test_suite.xlsx"):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = openpyxl.Workbook()

    # ── TestSuite Sheet ──────────────────────────────────────────
    ws = wb.active
    ws.title = "TestSuite"

    headers = [
        "TestCaseID", "TestName", "Enabled", "StepNo", "Keyword",
        "Locator", "Strategy", "Value", "Timeout", "DataSource", "Description"
    ]

    hdr_fill   = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    hdr_font   = Font(color="FFFFFF", bold=True, size=11)
    pass_fill  = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fail_fill  = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    dis_fill   = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    alt_fill   = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
    thin       = Side(style="thin", color="DDDDDD")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center")
    left       = Alignment(horizontal="left",   vertical="center")

    # Headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[1].height = 22

    # Sample rows
    rows = [
        # TC001 — Google search (enabled)
        ["TC001", "Google_Search_Test", "YES", 1, "NAVIGATE",       "",                   "css",   "https://www.google.com", 10, "",     "Open Google"],
        ["TC001", "Google_Search_Test", "YES", 2, "WAIT_FOR",       "input[name=q]",      "css",   "",                       10, "",     "Wait for search input"],
        ["TC001", "Google_Search_Test", "YES", 3, "TYPE",           "input[name=q]",      "css",   "KWAF automation",         5, "",     "Type search query"],
        ["TC001", "Google_Search_Test", "YES", 4, "SCREENSHOT",     "",                   "css",   "google_typed.png",        5, "",     "Capture typed state"],
        ["TC001", "Google_Search_Test", "YES", 5, "ASSERT_VISIBLE", "input[name=q]",      "css",   "",                        5, "",     "Assert search box visible"],

        # TC002 — Example.com heading check (enabled)
        ["TC002", "Example_Title_Test", "YES", 1, "NAVIGATE",       "",                   "css",   "https://example.com",    10, "",     "Open Example.com"],
        ["TC002", "Example_Title_Test", "YES", 2, "ASSERT_TEXT",    "h1",                 "css",   "Example Domain",          5, "",     "Verify page heading"],
        ["TC002", "Example_Title_Test", "YES", 3, "ASSERT_URL",     "",                   "css",   "example.com",             5, "",     "Verify URL"],
        ["TC002", "Example_Title_Test", "YES", 4, "SCREENSHOT",     "",                   "css",   "example_result.png",      5, "",     "Capture result"],

        # TC003 — Disabled test
        ["TC003", "Disabled_Test",      "NO",  1, "NAVIGATE",       "",                   "css",   "https://example.com",    10, "",     "Disabled — will not run"],
    ]

    fill_map = {
        "TC001": pass_fill,
        "TC002": alt_fill,
        "TC003": dis_fill,
    }

    for row_idx, row_data in enumerate(rows, start=2):
        tc_id = row_data[0]
        fill  = fill_map.get(tc_id, alt_fill)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = fill
            cell.alignment = center if col_idx in (3, 4, 7, 9) else left
            cell.border    = border
        ws.row_dimensions[row_idx].height = 18

    # Column widths
    widths = [12, 24, 8, 7, 18, 32, 10, 36, 8, 14, 30]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"

    # ── TestData Sheet ───────────────────────────────────────────
    ws2 = wb.create_sheet("TestData")
    data_headers = ["RowID", "username", "password", "email", "phone", "expected_text"]
    for col_idx, header in enumerate(data_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = border

    sample_data = [
        ["row1", "testuser1", "Pass@123",  "test1@example.com", "01711111111", "Welcome"],
        ["row2", "testuser2", "Pass@456",  "test2@example.com", "01722222222", "Dashboard"],
        ["row3", "adminuser", "Admin@789", "admin@example.com", "01733333333", "Admin Panel"],
    ]
    for row_idx, row_data in enumerate(sample_data, start=2):
        fill = pass_fill if row_idx % 2 == 0 else alt_fill
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = fill
            cell.border = border

    data_widths = [10, 14, 14, 24, 14, 18]
    for col_idx, w in enumerate(data_widths, start=1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

    ws2.freeze_panes = "A2"

    wb.save(output_path)
    print(f"\n  Template created: {output_path}")
    print(f"  Sheets: TestSuite ({len(rows)} sample steps), TestData ({len(sample_data)} rows)")
    print(f"\n  Usage:")
    print(f"    python run.py --list")
    print(f"    python run.py --headless\n")


if __name__ == "__main__":
    create_template()
