"""
Excel Data Provider — reads/writes test data from .xlsx files.
Each sheet = a data table. Row 1 = headers. Data starts row 2.
"""
from typing import Any, Optional, List
from loguru import logger
import openpyxl

from core.data.base_provider import BaseDataProvider


class ExcelProvider(BaseDataProvider):

    def __init__(self, filepath: str):
        self.filepath = filepath
        self._wb = None
        self._data_cache: dict = {}

    def connect(self) -> None:
        logger.info(f"[ExcelProvider] Loading: {self.filepath}")
        self._wb = openpyxl.load_workbook(self.filepath, data_only=True)
        self._preload_sheets()
        logger.success(f"[ExcelProvider] Loaded sheets: {list(self._data_cache.keys())}")

    def disconnect(self) -> None:
        if self._wb:
            self._wb.close()
        self._data_cache.clear()

    def _preload_sheets(self) -> None:
        for sheet_name in self._wb.sheetnames:
            ws = self._wb[sheet_name]
            headers = []
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c).strip() if c else f"col_{j}" for j, c in enumerate(row)]
                else:
                    if all(c is None for c in row):
                        continue
                    row_dict = {headers[j]: row[j] for j in range(min(len(headers), len(row)))}
                    rows.append(row_dict)
            self._data_cache[sheet_name] = rows

    def get_row(self, sheet_or_table: str, row_index: int) -> dict:
        rows = self._get_sheet_rows(sheet_or_table)
        if row_index >= len(rows):
            raise IndexError(
                f"[ExcelProvider] Row {row_index} out of range in '{sheet_or_table}' "
                f"(total rows: {len(rows)})"
            )
        return rows[row_index]

    def get_all_rows(self, sheet_or_table: str) -> list:
        return list(self._get_sheet_rows(sheet_or_table))

    def get_value(self, sheet_or_table: str, row_index: int, column: str) -> Any:
        return self.get_row(sheet_or_table, row_index).get(column)

    def write_result(self, sheet_or_table: str, row_index: int,
                     column: str, value: Any) -> None:
        wb = openpyxl.load_workbook(self.filepath)
        if sheet_or_table not in wb.sheetnames:
            raise ValueError(f"[ExcelProvider] Sheet '{sheet_or_table}' not found.")
        ws = wb[sheet_or_table]
        headers = [cell.value for cell in ws[1]]
        if column not in headers:
            raise ValueError(f"[ExcelProvider] Column '{column}' not found.")
        col_idx = headers.index(column) + 1
        ws.cell(row=row_index + 2, column=col_idx, value=value)
        wb.save(self.filepath)
        logger.debug(f"[ExcelProvider] Wrote '{value}' -> {sheet_or_table}[{row_index}][{column}]")

    def find_row(self, sheet_or_table: str, column: str, value: Any) -> Optional[dict]:
        for row in self._get_sheet_rows(sheet_or_table):
            if str(row.get(column, "")).strip() == str(value).strip():
                return row
        return None

    def _get_sheet_rows(self, sheet_name: str) -> List[dict]:
        if sheet_name not in self._data_cache:
            raise ValueError(
                f"[ExcelProvider] Sheet '{sheet_name}' not found. "
                f"Available: {list(self._data_cache.keys())}"
            )
        return self._data_cache[sheet_name]
