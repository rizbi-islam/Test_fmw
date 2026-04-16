"""
Site Inspector — standalone field discovery engine.
Visits any URL, extracts all interactive elements,
and generates a ready-to-use test case skeleton.

Usage:
    from core.inspector.site_inspector import SiteInspector
    inspector = SiteInspector(driver_type="playwright", headless=True)
    fields    = inspector.inspect("https://example.com/login")
    skeleton  = inspector.generate_skeleton("TC001", "Login_Test", fields)
    inspector.export_to_excel(skeleton, "assets/templates/tc001_skeleton.xlsx")
"""
import os
import time
from dataclasses import dataclass, field
from typing import List, Optional
from loguru import logger


@dataclass
class DiscoveredField:
    """Represents one discovered interactive element on a page."""
    tag:         str = ""
    field_type:  str = ""
    element_id:  str = ""
    name:        str = ""
    placeholder: str = ""
    label:       str = ""
    locator:     str = ""
    suggested_keyword: str = ""
    suggested_value:   str = ""


@dataclass
class TestStepSkeleton:
    """One row in the generated test case skeleton."""
    test_case_id:   str
    test_name:      str
    enabled:        str
    step_no:        int
    keyword:        str
    locator:        str
    strategy:       str
    value:          str
    timeout:        int
    description:    str


class SiteInspector:
    """
    Visits a URL, discovers all interactive fields,
    and generates a keyword-driven test case skeleton.
    """

    # Keyword suggestions per field type
    KEYWORD_MAP = {
        "password":  "TYPE",
        "email":     "TYPE",
        "text":      "TYPE",
        "tel":       "TYPE",
        "number":    "TYPE",
        "search":    "TYPE",
        "url":       "TYPE",
        "date":      "TYPE",
        "textarea":  "TYPE",
        "select":    "SELECT",
        "checkbox":  "CLICK",
        "radio":     "CLICK",
        "submit":    "CLICK",
        "button":    "CLICK",
        "reset":     "CLICK",
        "file":      "TYPE",
    }

    def __init__(self, driver_type: str = "playwright", headless: bool = True,
                 config: dict = None):
        self.driver_type = driver_type
        self.headless    = headless
        self.config      = config or {}
        self._driver     = None

    def inspect(self, url: str, wait_seconds: int = 2) -> List[DiscoveredField]:
        """
        Visit URL and return list of discovered fields.
        Args:
            url:          page to inspect
            wait_seconds: seconds to wait for page to render
        """
        logger.info(f"[SiteInspector] Inspecting: {url}")
        fields = []
        try:
            self._start_driver()
            self._driver.navigate(url)
            time.sleep(wait_seconds)
            raw_fields = self._extract_fields()
            fields = self._process_fields(raw_fields)
            logger.success(f"[SiteInspector] Found {len(fields)} field(s)")
        except Exception as e:
            logger.error(f"[SiteInspector] Inspection failed: {e}")
        finally:
            self._stop_driver()
        return fields

    def generate_skeleton(self, tc_id: str, tc_name: str, url: str,
                           fields: List[DiscoveredField]) -> List[TestStepSkeleton]:
        """Generate test step skeleton from discovered fields."""
        steps  = []
        step_n = 1

        # Step 1: Navigate
        steps.append(TestStepSkeleton(
            test_case_id=tc_id, test_name=tc_name, enabled="YES",
            step_no=step_n, keyword="NAVIGATE", locator="", strategy="css",
            value=url, timeout=10, description="Open page",
        ))
        step_n += 1

        for f in fields:
            if not f.locator:
                continue
            label = f.label or f.placeholder or f.name or f.field_type or "field"
            var   = "{" + label.lower().replace(" ", "_").replace("-", "_") + "}"

            if f.suggested_keyword == "TYPE":
                steps.append(TestStepSkeleton(
                    test_case_id=tc_id, test_name=tc_name, enabled="YES",
                    step_no=step_n, keyword="TYPE", locator=f.locator,
                    strategy="css", value=var, timeout=5,
                    description=f"Fill {label}",
                ))
            elif f.suggested_keyword == "SELECT":
                steps.append(TestStepSkeleton(
                    test_case_id=tc_id, test_name=tc_name, enabled="YES",
                    step_no=step_n, keyword="SELECT", locator=f.locator,
                    strategy="css", value="{option}", timeout=5,
                    description=f"Select {label}",
                ))
            elif f.suggested_keyword == "CLICK":
                if f.field_type in ("submit", "button"):
                    steps.append(TestStepSkeleton(
                        test_case_id=tc_id, test_name=tc_name, enabled="YES",
                        step_no=step_n, keyword="CLICK", locator=f.locator,
                        strategy="css", value="", timeout=5,
                        description=f"Click {label}",
                    ))
                else:
                    steps.append(TestStepSkeleton(
                        test_case_id=tc_id, test_name=tc_name, enabled="YES",
                        step_no=step_n, keyword="CLICK", locator=f.locator,
                        strategy="css", value="", timeout=5,
                        description=f"Click {label}",
                    ))
            step_n += 1

        # Final screenshot
        steps.append(TestStepSkeleton(
            test_case_id=tc_id, test_name=tc_name, enabled="YES",
            step_no=step_n, keyword="SCREENSHOT", locator="", strategy="css",
            value=f"{tc_id}_result.png", timeout=5, description="Capture result",
        ))
        return steps

    def export_to_excel(self, steps: List[TestStepSkeleton],
                         output_path: str) -> str:
        """Export skeleton steps to an Excel file ready to use in KWAF."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSuite"

        headers = ["TestCaseID","TestName","Enabled","StepNo","Keyword",
                   "Locator","Strategy","Value","Timeout","Description"]
        hf = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        hfont = Font(color="FFFFFF", bold=True)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hf; c.font = hfont
            c.alignment = Alignment(horizontal="center")

        for ri, step in enumerate(steps, 2):
            row = [
                step.test_case_id, step.test_name, step.enabled,
                step.step_no, step.keyword, step.locator, step.strategy,
                step.value, step.timeout, step.description,
            ]
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)

        col_widths = [12,24,8,7,18,30,10,30,8,30]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        wb.save(output_path)
        logger.success(f"[SiteInspector] Skeleton saved -> {output_path}")
        return output_path

    def export_to_csv(self, steps: List[TestStepSkeleton], output_path: str) -> str:
        """Export skeleton steps as CSV."""
        import csv
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["TestCaseID","TestName","Enabled","StepNo","Keyword",
                              "Locator","Strategy","Value","Timeout","Description"])
            for step in steps:
                writer.writerow([
                    step.test_case_id, step.test_name, step.enabled,
                    step.step_no, step.keyword, step.locator, step.strategy,
                    step.value, step.timeout, step.description,
                ])
        logger.success(f"[SiteInspector] CSV saved -> {output_path}")
        return output_path

    # ── Private helpers ───────────────────────────────────────

    def _start_driver(self) -> None:
        from core.drivers.driver_factory import DriverFactory
        driver_config = {**self.config.get("driver", {}),
                         "headless": self.headless}
        self._driver = DriverFactory.get(self.driver_type, driver_config)
        self._driver.start()

    def _stop_driver(self) -> None:
        if self._driver:
            try: self._driver.quit()
            except Exception: pass
            self._driver = None

    def _extract_fields(self) -> list:
        """Run JavaScript on page to extract all interactive elements."""
        script = """
        var results = [];
        var selectors = 'input, select, textarea, button';
        document.querySelectorAll(selectors).forEach(function(el) {
            if (el.type === 'hidden') return;
            var label = '';
            if (el.id) {
                var lbl = document.querySelector('label[for="' + el.id + '"]');
                if (lbl) label = lbl.innerText.trim();
            }
            if (!label && el.closest('label')) {
                label = el.closest('label').innerText.replace(el.value||'','').trim();
            }
            results.push({
                tag:         el.tagName.toLowerCase(),
                type:        el.type || '',
                id:          el.id || '',
                name:        el.name || '',
                placeholder: el.placeholder || '',
                label:       label,
                locator:     el.id
                               ? '#' + el.id
                               : (el.name ? '[name="' + el.name + '"]' : ''),
                visible:     el.offsetParent !== null,
            });
        });
        return results;
        """
        return self._driver.execute_script(script) or []

    def _process_fields(self, raw: list) -> List[DiscoveredField]:
        fields = []
        seen   = set()
        for item in raw:
            locator = item.get("locator", "")
            if not locator or locator in seen:
                continue
            if not item.get("visible", True):
                continue
            seen.add(locator)
            tag   = item.get("tag", "")
            ftype = item.get("type", "")
            kw    = self.KEYWORD_MAP.get(
                ftype if tag != "select" else "select",
                "CLICK"
            )
            label = item.get("label","") or item.get("placeholder","") or item.get("name","")
            var   = "{" + label.lower().replace(" ","_")[:20] + "}" if label else "{value}"
            fields.append(DiscoveredField(
                tag=tag, field_type=ftype,
                element_id=item.get("id",""), name=item.get("name",""),
                placeholder=item.get("placeholder",""), label=label,
                locator=locator, suggested_keyword=kw, suggested_value=var,
            ))
        return fields
