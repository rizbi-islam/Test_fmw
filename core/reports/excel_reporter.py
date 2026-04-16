"""
Excel Result Reporter.
Writes test execution results back into the Excel test suite file.
Adds a "Results" sheet with pass/fail per test case and per step.
Also writes result status into the original TestSuite sheet (Result column).
"""
import os
from datetime import datetime
from loguru import logger
from core.reports.base_reporter import BaseReporter


class ExcelReporter(BaseReporter):

    def __init__(self, config: dict):
        super().__init__(config)
        self.suite_path = config.get(
            "suite_path", "assets/templates/test_suite.xlsx"
        )

    def generate(self, suite_result) -> str:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        os.makedirs(self.output_dir, exist_ok=True)
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe     = suite_result.suite_name.replace(" ", "_").replace("/", "_")
        filepath = os.path.join(self.output_dir, f"results_{safe}_{ts}.xlsx")

        wb = openpyxl.Workbook()

        # ── Styles ────────────────────────────────────────────
        hdr_fill  = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        hdr_font  = Font(color="FFFFFF", bold=True, size=11)
        pass_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        fail_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        skip_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
        alt_fill  = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
        thin      = openpyxl.styles.Side(style="thin", color="DDDDDD")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)
        center    = Alignment(horizontal="center", vertical="center")
        left      = Alignment(horizontal="left",   vertical="center")

        def status_fill(status):
            return pass_fill if status=="PASS" else (
                   fail_fill if status=="FAIL" else skip_fill)

        # ── Sheet 1: Suite Summary ────────────────────────────
        ws1 = wb.active
        ws1.title = "Suite Summary"

        summary_headers = ["TestCaseID","TestName","Status","TotalSteps",
                           "PassedSteps","FailedSteps","Duration(s)","RunAt"]
        for ci, h in enumerate(summary_headers, 1):
            c = ws1.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = center; c.border = border
        ws1.row_dimensions[1].height = 20

        for ri, result in enumerate(suite_result.results, 2):
            row_data = [
                result.test_id, result.test_name, result.status,
                result.total_steps, result.passed_steps, result.failed_steps,
                result.elapsed_seconds,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ]
            sf = status_fill(result.status)
            for ci, val in enumerate(row_data, 1):
                c = ws1.cell(row=ri, column=ci, value=val)
                c.fill      = sf if ci == 3 else (pass_fill if ri%2==0 else alt_fill)
                c.alignment = center if ci in (3,4,5,6,7) else left
                c.border    = border

        col_widths_1 = [14, 28, 10, 12, 12, 12, 12, 20]
        for ci, w in enumerate(col_widths_1, 1):
            ws1.column_dimensions[
                openpyxl.utils.get_column_letter(ci)].width = w

        ws1.freeze_panes = "A2"

        # ── Sheet 2: Step Details ─────────────────────────────
        ws2 = wb.create_sheet("Step Details")

        step_headers = ["TestCaseID","TestName","StepNo","Keyword",
                        "Status","Message","Duration(ms)"]
        for ci, h in enumerate(step_headers, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = center; c.border = border
        ws2.row_dimensions[1].height = 20

        row = 2
        for result in suite_result.results:
            for step in result.step_results:
                d = step.to_dict() if hasattr(step,"to_dict") else vars(step)
                sf = status_fill(str(d.get("status","")))
                row_data = [
                    result.test_id,
                    result.test_name,
                    d.get("step_no",""),
                    d.get("keyword",""),
                    d.get("status",""),
                    str(d.get("message",""))[:200],
                    round(float(d.get("elapsed_ms",0)),2),
                ]
                for ci, val in enumerate(row_data, 1):
                    c = ws2.cell(row=row, column=ci, value=val)
                    c.fill      = sf if ci==5 else (pass_fill if row%2==0 else alt_fill)
                    c.alignment = center if ci in (3,5,7) else left
                    c.border    = border
                row += 1

        col_widths_2 = [14, 26, 8, 18, 10, 60, 14]
        for ci, w in enumerate(col_widths_2, 1):
            ws2.column_dimensions[
                openpyxl.utils.get_column_letter(ci)].width = w

        ws2.freeze_panes = "A2"

        # ── Sheet 3: Run Metadata ─────────────────────────────
        ws3 = wb.create_sheet("Run Metadata")
        meta = [
            ("Suite Name",    suite_result.suite_name),
            ("Run Date",      datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Tests",   suite_result.total),
            ("Passed",        suite_result.passed),
            ("Failed",        suite_result.failed),
            ("Skipped",       suite_result.skipped),
            ("Pass Rate",     f"{int(suite_result.passed/suite_result.total*100) if suite_result.total else 0}%"),
            ("Framework",     "KWAF v2.0"),
        ]
        for ri, (key, val) in enumerate(meta, 1):
            ws3.cell(row=ri, column=1, value=key).font  = Font(bold=True)
            ws3.cell(row=ri, column=2, value=val)
        ws3.column_dimensions["A"].width = 20
        ws3.column_dimensions["B"].width = 30

        wb.save(filepath)
        logger.success(f"[ExcelReporter] Report -> {filepath}")
        return filepath
