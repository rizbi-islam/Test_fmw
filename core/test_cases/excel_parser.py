"""
Excel Test Suite Parser.
Reads test cases from a structured .xlsx file.

Sheet: "TestSuite"
Columns: TestCaseID | TestName | Enabled | StepNo | Keyword |
         Locator | Strategy | Value | Timeout | DataSource | Description
"""
from typing import List, Dict
from loguru import logger
import openpyxl

from core.test_cases.test_case_model import TestCase, TestStep, TestSuite

REQUIRED_COLUMNS = [
    "TestCaseID", "TestName", "Enabled", "StepNo", "Keyword", "Locator", "Strategy", "Value"
]


class ExcelParser:
    """Parses .xlsx files into TestSuite objects."""

    def __init__(self, filepath: str, sheet_name: str = "TestSuite"):
        self.filepath = filepath
        self.sheet_name = sheet_name

    def parse(self) -> TestSuite:
        logger.info(f"[ExcelParser] Parsing: {self.filepath} | Sheet: {self.sheet_name}")
        wb = openpyxl.load_workbook(self.filepath, data_only=True)

        if self.sheet_name not in wb.sheetnames:
            raise ValueError(
                f"[ExcelParser] Sheet '{self.sheet_name}' not found. "
                f"Available: {wb.sheetnames}"
            )

        ws = wb[self.sheet_name]
        headers = self._get_headers(ws)
        self._validate_headers(headers)
        rows = self._read_rows(ws, headers)
        suite = self._build_suite(rows)

        logger.success(
            f"[ExcelParser] Parsed {suite.total} test cases ({suite.enabled_count} enabled)"
        )
        return suite

    def _get_headers(self, ws) -> Dict[str, int]:
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
        return headers

    def _validate_headers(self, headers: dict) -> None:
        missing = [col for col in REQUIRED_COLUMNS if col not in headers]
        if missing:
            raise ValueError(
                f"[ExcelParser] Missing required columns: {missing}. Found: {list(headers.keys())}"
            )

    def _read_rows(self, ws, headers: dict) -> List[dict]:
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
            row_dict = {}
            for col_name, col_idx in headers.items():
                row_dict[col_name] = row[col_idx - 1]
            rows.append(row_dict)
        return rows

    def _build_suite(self, rows: List[dict]) -> TestSuite:
        suite = TestSuite(suite_id="excel_suite", name="Excel Test Suite")
        test_cases: Dict[str, TestCase] = {}

        for row in rows:
            tc_id = str(row.get("TestCaseID", "")).strip()
            if not tc_id:
                continue

            if tc_id not in test_cases:
                enabled_raw = str(row.get("Enabled", "YES")).upper()
                enabled = enabled_raw in ("YES", "TRUE", "1", "Y")
                tc = TestCase(
                    test_id=tc_id,
                    name=str(row.get("TestName", tc_id)).strip(),
                    enabled=enabled,
                    data_source=str(row.get("DataSource", "excel")).strip(),
                    description=str(row.get("Description", "")).strip(),
                )
                test_cases[tc_id] = tc
                suite.add_test_case(tc)

            step_no = row.get("StepNo")
            keyword = str(row.get("Keyword", "")).strip().upper()
            if not keyword or not step_no:
                continue

            step = TestStep(
                step_no=int(step_no),
                keyword=keyword,
                locator=str(row.get("Locator", "") or "").strip(),
                strategy=str(row.get("Strategy", "css") or "css").strip().lower(),
                value=str(row.get("Value", "") or "").strip(),
                timeout=int(row.get("Timeout", 10) or 10),
                enabled=True,
                data_source=str(row.get("DataSource", "") or "").strip(),
                description=str(row.get("StepDescription", "") or "").strip(),
            )
            test_cases[tc_id].add_step(step)

        return suite
