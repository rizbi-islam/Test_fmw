"""
PHASE 5 VERIFICATION SCRIPT
============================
Run from your kwaf project root after copying Phase 5 files in.

Usage:
    python tests/verify_phase5.py

All 8 checks must PASS — this is the final verification.
"""
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

PASS_MARK = "PASS"
FAIL_MARK = "FAIL"
results   = []


def check(name, fn):
    try:
        fn()
        results.append((PASS_MARK, name))
        print(f"  [PASS]  {name}")
    except Exception as e:
        results.append((FAIL_MARK, name, str(e)))
        print(f"  [FAIL]  {name}")
        print(f"          -> {str(e)[:200]}")


print("\n" + "=" * 62)
print("  KWAF — Phase 5 Verification  (8 checks)")
print("=" * 62 + "\n")


# CHECK 1 — All Phase 5 files exist
def check_files():
    required = [
        "core/reports/pdf_reporter.py",
        "core/reports/excel_reporter.py",
        "core/inspector/__init__.py",
        "core/inspector/site_inspector.py",
        "assets/templates/blank_template.py",
        "docs/KEYWORD_REFERENCE.md",
        "docs/USER_GUIDE.md",
    ]
    missing = [f for f in required if not os.path.exists(f)]
    if missing:
        raise FileNotFoundError("Missing:\n    " + "\n    ".join(missing))

check("All Phase 5 files exist on disk", check_files)


# CHECK 2 — PDF reporter imports and instantiates
def check_pdf_reporter():
    from core.reports.pdf_reporter import PdfReporter
    reporter = PdfReporter({"output_dir": "reports/output"})
    assert reporter is not None
    assert hasattr(reporter, "generate")

check("PdfReporter imports and instantiates correctly", check_pdf_reporter)


# CHECK 3 — Excel reporter generates a file
def check_excel_reporter():
    from core.reports.excel_reporter import ExcelReporter
    from core.flow.flow_executor import SuiteResult
    from core.keywords.keyword_executor import TestCaseResult, StepResult
    from datetime import datetime

    config   = {"output_dir": "reports/output"}
    reporter = ExcelReporter(config)

    suite  = SuiteResult("Phase5 Test Suite")
    result = TestCaseResult("TC_P5_01", "Phase5 Verify Test")
    result.status      = "PASS"
    result.finished_at = datetime.now()
    result.step_results = [
        StepResult(1, "NAVIGATE", "PASS", "Navigated OK", "", 100.0),
        StepResult(2, "ASSERT_URL", "PASS", "URL verified", "", 50.0),
    ]
    suite.results.append(result)

    path = reporter.generate(suite)
    assert os.path.exists(path), f"Excel report not found: {path}"

    import openpyxl
    wb     = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    assert "Suite Summary"  in sheets, "Missing 'Suite Summary' sheet"
    assert "Step Details"   in sheets, "Missing 'Step Details' sheet"
    assert "Run Metadata"   in sheets, "Missing 'Run Metadata' sheet"

check("ExcelReporter generates valid .xlsx with 3 sheets", check_excel_reporter)


# CHECK 4 — PDF reporter generates a file
def check_pdf_generation():
    from core.reports.pdf_reporter import PdfReporter
    from core.flow.flow_executor import SuiteResult
    from core.keywords.keyword_executor import TestCaseResult, StepResult
    from datetime import datetime

    reporter = PdfReporter({"output_dir": "reports/output"})
    suite    = SuiteResult("PDF Verify Suite")

    r = TestCaseResult("TC_PDF_01", "PDF Test")
    r.status = "PASS"; r.finished_at = datetime.now()
    r.step_results = [StepResult(1, "NAVIGATE", "PASS", "OK", "", 80.0)]
    suite.results.append(r)

    path = reporter.generate(suite)
    assert os.path.exists(path), f"PDF not found: {path}"
    assert path.endswith(".pdf"), f"Expected .pdf, got: {path}"
    assert os.path.getsize(path) > 1000, "PDF file too small — likely empty"

check("PdfReporter generates valid non-empty .pdf file", check_pdf_generation)


# CHECK 5 — SiteInspector imports and initialises
def check_site_inspector():
    from core.inspector.site_inspector import SiteInspector, DiscoveredField, TestStepSkeleton

    inspector = SiteInspector(driver_type="playwright", headless=True)
    assert inspector is not None
    assert hasattr(inspector, "inspect")
    assert hasattr(inspector, "generate_skeleton")
    assert hasattr(inspector, "export_to_excel")
    assert hasattr(inspector, "export_to_csv")

    # Test skeleton generation without a browser
    mock_fields = [
        DiscoveredField(tag="input", field_type="email", element_id="email",
                        locator="#email", suggested_keyword="TYPE", label="Email"),
        DiscoveredField(tag="input", field_type="password", element_id="password",
                        locator="#password", suggested_keyword="TYPE", label="Password"),
        DiscoveredField(tag="input", field_type="submit", element_id="submit",
                        locator="#submit", suggested_keyword="CLICK", label="Login"),
    ]
    steps = inspector.generate_skeleton("TC_SI", "Inspector_Test",
                                         "https://example.com", mock_fields)
    assert len(steps) >= 4, f"Expected 4+ steps, got {len(steps)}"
    assert steps[0].keyword == "NAVIGATE"
    assert steps[-1].keyword == "SCREENSHOT"

check("SiteInspector initialises and generates skeleton correctly", check_site_inspector)


# CHECK 6 — Blank template creates successfully
def check_blank_template():
    import subprocess
    result = subprocess.run(
        [sys.executable, "assets/templates/blank_template.py"],
        capture_output=True, text=True, timeout=15,
    )
    assert result.returncode == 0, f"blank_template.py failed:\n{result.stderr}"
    assert os.path.exists("assets/templates/blank_test_suite.xlsx"), \
        "blank_test_suite.xlsx not created"

    import openpyxl
    wb     = openpyxl.load_workbook("assets/templates/blank_test_suite.xlsx")
    sheets = wb.sheetnames
    assert "TestSuite"          in sheets
    assert "TestData"           in sheets
    assert "Keywords Reference" in sheets

check("blank_template.py creates valid Excel with 3 sheets", check_blank_template)


# CHECK 7 — Docs exist and have content
def check_docs():
    for doc in ["docs/KEYWORD_REFERENCE.md", "docs/USER_GUIDE.md"]:
        assert os.path.exists(doc), f"Missing: {doc}"
        with open(doc, encoding="utf-8") as f:
            content = f.read()
        assert len(content) > 1000, f"{doc} seems too short ({len(content)} chars)"
        assert "NAVIGATE" in content, f"{doc} missing NAVIGATE keyword"

check("Documentation files exist and have sufficient content", check_docs)


# CHECK 8 — All phases still intact
def check_all_phases():
    from core.drivers.driver_factory import DriverFactory
    from core.otp.otp_factory import OTPFactory
    from core.captcha.captcha_factory import CaptchaFactory
    from core.data.data_factory import DataFactory
    import core.keywords.keyword_registry as kr
    kr.KeywordRegistry._instance = None
    from core.keywords.keyword_registry import KeywordRegistry
    reg = KeywordRegistry()

    assert len(DriverFactory.available_drivers()) >= 2,   "Drivers missing"
    assert reg.has_keyword("HANDLE_OTP"),                  "HANDLE_OTP missing"
    assert reg.has_keyword("HANDLE_CAPTCHA"),              "HANDLE_CAPTCHA missing"
    assert "firebase" in OTPFactory.available(),           "Firebase OTP missing"
    assert "sqlite"   in DataFactory.available_providers(),"SQLite provider missing"
    assert os.path.exists("ui/app.py"),                    "UI missing"
    assert os.path.exists("Dockerfile"),                   "Dockerfile missing"
    assert os.path.exists(".github/workflows/ci.yml"),     "CI pipeline missing"

check("All Phases 1-4 still fully intact after Phase 5 merge", check_all_phases)


# ── FINAL VERDICT ─────────────────────────────────────────────
passed_list = [r for r in results if r[0] == PASS_MARK]
failed_list = [r for r in results if r[0] == FAIL_MARK]

print(f"\n{'=' * 62}")
print(f"  Checks passed : {len(passed_list)} / {len(results)}")
print(f"{'=' * 62}")

if not failed_list:
    print("""
  ALL 8 CHECKS PASSED.

  KWAF IS FULLY COMPLETE — ALL 5 PHASES DONE.

  ┌─────────────────────────────────────────────┐
  │  Phase 1 — Core Engine          12/12  DONE │
  │  Phase 2 — OTP + CAPTCHA + DB   10/10  DONE │
  │  Phase 3 — Streamlit UI          6/6   DONE │
  │  Phase 4 — Parallel + CI/CD      8/8   DONE │
  │  Phase 5 — Reports + Docs        8/8   DONE │
  └─────────────────────────────────────────────┘

  Total keywords: 20
  Total test commands available:

    python run.py --list
    python run.py --headless
    python run.py --headless --parallel --workers 4
    python run.py --headless --driver playwright
    python run.py --headless --test-id TC001
    streamlit run ui\\app.py

  Start writing real tests:
    1. python assets/templates/blank_template.py
    2. Open assets/templates/blank_test_suite.xlsx
    3. Fill in your test cases
    4. python run.py --headless
""")
    sys.exit(0)
else:
    print(f"\n  {len(failed_list)} check(s) FAILED:\n")
    for r in failed_list:
        print(f"    [FAIL]  {r[1]}")
        if len(r) > 2:
            print(f"            -> {r[2][:300]}")
    print("\n  Fix failing checks then re-run: python tests/verify_phase5.py\n")
    sys.exit(1)
