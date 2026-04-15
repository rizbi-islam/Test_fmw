import ui._pathfix  # noqa
import time
import streamlit as st
from ui.state import init_state, get_config


def render():
    init_state()
    config = get_config()

    st.title("🔍 Site Inspector")
    st.caption("Visit any site and auto-discover all form fields and interactive elements.")
    st.divider()

    col1, col2 = st.columns([4, 1])
    url         = col1.text_input("Target URL", placeholder="https://example.com/login",
                                   value=st.session_state.get("inspect_url", ""))
    driver_type = col2.selectbox("Driver", ["selenium", "playwright"])
    headless    = st.checkbox("Run headless (no browser window)", value=True)

    if st.button("🔍 Inspect Site", use_container_width=True, type="primary"):
        if not url:
            st.warning("Please enter a URL.")
            return
        st.session_state.inspect_url = url
        with st.spinner(f"Inspecting {url} ..."):
            fields = _inspect_site(url, driver_type, headless, config)
            st.session_state.inspected_fields = fields

    fields = st.session_state.get("inspected_fields", [])

    if fields:
        st.divider()
        st.success(f"Found {len(fields)} field(s) on the page.")
        import pandas as pd
        st.dataframe(pd.DataFrame(fields), use_container_width=True)
        st.divider()

        st.subheader("Generated Test Case Skeleton")
        tc_id   = st.text_input("Test Case ID",   value="TC_NEW")
        tc_name = st.text_input("Test Case Name", value="New_Test")

        skeleton_rows = _generate_skeleton(tc_id, tc_name, url, fields)
        sk_df = pd.DataFrame(skeleton_rows, columns=[
            "TestCaseID", "TestName", "Enabled", "StepNo",
            "Keyword", "Locator", "Strategy", "Value", "Timeout", "Description"
        ])
        st.dataframe(sk_df, use_container_width=True)

        csv = sk_df.to_csv(index=False)
        st.download_button(
            label="⬇️ Download Skeleton as CSV",
            data=csv, file_name=f"{tc_id}_skeleton.csv",
            mime="text/csv", use_container_width=True,
        )

        try:
            import io, openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "TestSuite"
            headers = list(sk_df.columns)
            hf = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            hfont = Font(color="FFFFFF", bold=True)
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.fill = hf; c.font = hfont
                c.alignment = Alignment(horizontal="center")
            for ri, row in enumerate(skeleton_rows, 2):
                for ci, val in enumerate(row, 1):
                    ws.cell(row=ri, column=ci, value=val)
            buf = io.BytesIO()
            wb.save(buf); buf.seek(0)
            st.download_button(
                label="⬇️ Download Skeleton as Excel",
                data=buf, file_name=f"{tc_id}_skeleton.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.warning(f"Excel export failed: {e}")


def _inspect_site(url, driver_type, headless, config):
    from core.drivers.driver_factory import DriverFactory
    driver_config = {**config.get("driver", {}), "headless": headless}
    driver = DriverFactory.get(driver_type, driver_config)
    fields = []
    try:
        driver.start()
        driver.navigate(url)
        time.sleep(2)
        script = """
        var results = [];
        document.querySelectorAll('input,select,textarea,button').forEach(function(el){
            if(el.type==='hidden') return;
            var label='';
            if(el.id){var l=document.querySelector('label[for="'+el.id+'"]');if(l)label=l.innerText.trim();}
            results.push({
                tag:el.tagName.toLowerCase(), type:el.type||'',
                id:el.id||'', name:el.name||'',
                placeholder:el.placeholder||'', label:label,
                locator:el.id?'#'+el.id:(el.name?'[name="'+el.name+'"]':'')
            });
        });
        return results;
        """
        raw = driver.execute_script(script)
        if raw:
            for item in raw:
                if item.get("type") in ("hidden", ""):
                    continue
                fields.append({
                    "Tag":      item.get("tag",""),
                    "Type":     item.get("type",""),
                    "ID":       item.get("id",""),
                    "Name":     item.get("name",""),
                    "Label":    item.get("label","") or item.get("placeholder",""),
                    "Locator":  item.get("locator",""),
                    "Keyword":  _suggest_keyword(item),
                })
    except Exception as e:
        st.error(f"Inspection failed: {e}")
    finally:
        try: driver.quit()
        except Exception: pass
    return fields


def _suggest_keyword(field):
    tag, ftype = field.get("tag",""), field.get("type","")
    if tag == "select": return "SELECT"
    if tag == "button" or ftype in ("submit","button","reset"): return "CLICK"
    if ftype in ("password","email","text","tel","number","search"): return "TYPE"
    if tag == "textarea": return "TYPE"
    return "CLICK"


def _generate_skeleton(tc_id, tc_name, url, fields):
    rows = []
    step = 1
    rows.append([tc_id, tc_name, "YES", step, "NAVIGATE", "", "css", url, 10, "Open page"])
    step += 1
    for field in fields:
        locator = field.get("Locator","")
        keyword = field.get("Keyword","TYPE")
        label   = field.get("Label","") or "field"
        if not locator: continue
        if keyword == "TYPE":
            rows.append([tc_id,tc_name,"YES",step,"TYPE",locator,"css",
                         f"{{{label.lower().replace(' ','_')}}}",5,f"Fill {label}"])
        elif keyword == "SELECT":
            rows.append([tc_id,tc_name,"YES",step,"SELECT",locator,"css","{option}",5,f"Select {label}"])
        elif keyword == "CLICK":
            rows.append([tc_id,tc_name,"YES",step,"CLICK",locator,"css","",5,f"Click {label}"])
        step += 1
    rows.append([tc_id,tc_name,"YES",step,"SCREENSHOT","","css",f"{tc_id}_result.png",5,"Capture result"])
    return rows
